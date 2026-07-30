import os
import json
import requests
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], 
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class PromptRequest(BaseModel):
    prompt: str

OLLAMA_CLOUD_URL = "https://ollama.com/api/chat"

def get_free_cloud_data(user_prompt: str) -> dict:
    api_key = os.environ.get("OLLAMA_API_KEY")
    if not api_key:
        raise Exception("System Error: Missing OLLAMA_API_KEY environment variable.")

    system_instruction = (
        "You are an expert data designer. Convert the user request into a highly professional JSON object matching this schema exactly:\n"
        "{\n"
        "  \"title\": \"Main Title of the Sheet\",\n"
        "  \"theme_color\": \"Hex code for headers (e.g., #1F4E78 for Navy)\",\n"
        "  \"text_color\": \"#FFFFFF\",\n"
        "  \"headers\": [\"Column 1\", \"Column 2\"],\n"
        "  \"rows\": [[\"Row1Col1\", \"Row1Col2\"]],\n"
        "  \"column_types\": [\"text\", \"number\"]\n"
        "}\n"
        "Output ONLY raw JSON. No markdown wrappers, no conversational filler."
    )
    
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "model": "gpt-oss:120b",
        "messages": [
            {"role": "user", "content": f"{system_instruction}\n\nCreate data for: {user_prompt}"}
        ],
        "stream": False
    }
    
    try:
        response = requests.post(OLLAMA_CLOUD_URL, headers=headers, json=payload)
        res_data = response.json()
        raw_text = res_data["message"]["content"].strip()
        
        if "```json" in raw_text:
            raw_text = raw_text.split("```json")[1].split("```")[0].strip()
        elif "```" in raw_text:
            raw_text = raw_text.split("```")[1].split("```")[0].strip()
            
        return json.loads(raw_text)
    except Exception as e:
        raise Exception(f"Free Cloud AI Pipeline Error: {str(e)}")

def build_excel_file(data: dict, filepath: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "AI Design Report"
    ws.views.sheetView[0].showGridLines = True
    
    brand_color = data.get("theme_color", "#1F4E78").lstrip('#')
    txt_color = data.get("text_color", "#FFFFFF").lstrip('#')
    
    title_font = Font(name="Segoe UI", size=16, bold=True, color=brand_color)
    header_font = Font(name="Segoe UI", size=11, bold=True, color=txt_color)
    header_fill = PatternFill(start_color=brand_color, end_color=brand_color, fill_type="solid")
    data_font = Font(name="Segoe UI", size=10)
    zebra_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    thin_side = Side(style="thin", color="E2E8F0")
    data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws['A1'] = data.get("title", "Report")
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 35
    
    headers = data.get("headers", [])
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = data_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 28

    rows = data.get("rows", [])
    column_types = data.get("column_types", ["text"] * len(headers))
    
    current_row = 4
    for row_idx, row_data in enumerate(rows):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            col_type = column_types[col_idx - 1] if col_idx - 1 < len(column_types) else "text"
            
            if col_type in ["number", "currency"] and str(value).replace('.','',1).replace('-','',1).isdigit():
                cell.value = float(value)
                cell.number_format = '$#,##0.00' if col_type == "currency" else '#,##0'
            else:
                cell.value = value
                
            cell.font = data_font
            cell.border = data_border
            if row_idx % 2 == 1:
                cell.fill = zebra_fill
        current_row += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col if cell.row != 1)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(filepath)

@app.post("/api/generate")
async def generate_sheet(payload: PromptRequest):
    filename = "free_cloud_output.xlsx"
    try:
        structured_data = get_free_cloud_data(payload.prompt)
        build_excel_file(structured_data, filename)
        return FileResponse(path=filename, filename="AI_Report.xlsx")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    import os
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)


